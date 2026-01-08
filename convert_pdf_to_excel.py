import pdfplumber
import openpyxl
import fitz  # PyMuPDF
import sys
import os
import re

def extract_images_from_pdf(pdf_file, output_folder):
    pdf_document = fitz.open(pdf_file)
    image_count = 0
    
    for page_number in range(len(pdf_document)):
        page = pdf_document.load_page(page_number)
        for img_index, img in enumerate(page.get_images(full=True)):
            xref = img[0]
            base_image = pdf_document.extract_image(xref)
            image_bytes = base_image["image"]
            image_filename = os.path.join(output_folder, f"page_{page_number + 1}_img_{img_index + 1}.png")
            with open(image_filename, "wb") as image_file:
                image_file.write(image_bytes)
            image_count += 1
    
    return image_count

def convert_pdf_to_excel(pdf_file, excel_file):
    try:
        output_dir = os.path.dirname(excel_file)
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        # Extract images first
        image_count = extract_images_from_pdf(pdf_file, output_dir)
        print(f"Extracted {image_count} images.")

        # Open the PDF file
        print(f"Opening PDF file: {pdf_file}")
        with pdfplumber.open(pdf_file) as pdf:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            headers = None

            # Loop through each page in the PDF
            for page_num, page in enumerate(pdf.pages, start=1):
                # Extract tables
                tables = page.extract_table()
                if tables:
                    if not headers:
                        headers = tables[0]
                        sheet.append(headers)  # Add headers to the first row
                        header_font = openpyxl.styles.Font(bold=True)
                        for col_num in range(1, len(headers) + 1):
                            cell = sheet.cell(row=1, column=col_num)
                            cell.font = header_font
                    
                    # Append rows below headers
                    for row_idx, row in enumerate(tables[1:], start=sheet.max_row + 1):  # Skip headers
                        for col_idx, cell in enumerate(row, start=1):
                            sheet.cell(row=row_idx, column=col_idx, value=cell)
                else:
                    # Extract text
                    text = page.extract_text()
                    if text:
                        lines = text.split('\n')
                        for line_num, line in enumerate(lines, start=sheet.max_row + 1):  # Append to next row
                            # Adjust the regex pattern based on your actual data format
                            columns = [col.strip() for col in re.split(r'(?<!\s)\s(?!\s)', line.strip())]
                            if headers and len(columns) == len(headers):
                                for col_num, column_value in enumerate(columns, start=1):
                                    sheet.cell(row=line_num, column=col_num, value=column_value)
                            elif not headers:
                                if len(columns) > 1:  # Heuristic to determine if data is likely in columns
                                    sheet.append(columns)

            # Save the Excel file
            print(f"Saving Excel file: {excel_file}")
            workbook.save(excel_file)
            print(f"Conversion complete. Excel file saved to {excel_file}")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python convert_pdf_to_excel.py <input_pdf_file> <output_excel_file>")
        sys.exit(1)

    pdf_file = sys.argv[1]
    excel_file = sys.argv[2]
    
    # Ensure paths are correctly formed
    pdf_file = os.path.abspath(pdf_file)
    excel_file = os.path.abspath(excel_file)
    
    if not os.path.isfile(pdf_file):
        print(f"Error: The input PDF file does not exist at {pdf_file}")
        sys.exit(1)
    
    convert_pdf_to_excel(pdf_file, excel_file)